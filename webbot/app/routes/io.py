"""IO-style converter: URL in -> JSON out.

Mirrors canada.ca's "IO" app concept: pull data from an external URL,
normalize it, and emit JSON ready for downstream systems (e.g. AEM).

Supported formats (v1):
  - .xlsx  (Excel 2007+, via openpyxl)
  - .csv / .tsv (delimiter auto-sniffed)

Not yet: .xls (legacy Excel 97-2003), JSON/XML feeds, pagination.

SSRF note: by default private/reserved IPs are rejected. Set env
IO_ALLOW_PRIVATE_URLS=1 to allow internal data sources.
"""
import csv
import io
import ipaddress
import os
import socket
import urllib.parse

import requests
from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel, Field

router = APIRouter(prefix="/api/v1/io", tags=["io"])

MAX_BYTES = 20 * 1024 * 1024  # 20 MB download cap
DEFAULT_MAX_ROWS = 10000
ALLOW_PRIVATE = os.environ.get("IO_ALLOW_PRIVATE_URLS", "0") == "1"


class ConvertRequest(BaseModel):
    url: str | None = None
    datasource: str | None = None  # alias for url (mustache template field name)
    sheet: str | None = None  # sheet name, or 1-based index as string like "2"
    header_row: int = Field(1, ge=1)  # 1-based row used as column headers
    max_rows: int = Field(DEFAULT_MAX_ROWS, ge=1, le=100000)

    def resolved_url(self) -> str:
        return self.url or self.datasource or ""


# ---------------------------------------------------------------- helpers

def _check_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    if parsed.scheme not in ("http", "https"):
        raise HTTPException(400, "Only http/https URLs are supported")
    if not parsed.netloc:
        raise HTTPException(400, "URL must include a host")
    if not ALLOW_PRIVATE:
        host = parsed.hostname
        try:
            infos = socket.getaddrinfo(host, None)
        except socket.gaierror:
            raise HTTPException(400, f"Cannot resolve host: {host}")
        for info in infos:
            ip = ipaddress.ip_address(info[4][0])
            if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_multicast or ip.is_reserved:
                raise HTTPException(400, f"Private/reserved IP not allowed: {ip}")
    return url


def _download(url: str) -> bytes:
    try:
        r = requests.get(url, timeout=15, allow_redirects=True, stream=True)
        r.raise_for_status()
    except requests.RequestException as e:
        raise HTTPException(502, f"Download failed: {e}")
    total = 0
    chunks = []
    for chunk in r.iter_content(64 * 1024):
        total += len(chunk)
        if total > MAX_BYTES:
            raise HTTPException(413, f"File too large (> {MAX_BYTES // 1024 // 1024} MB)")
        chunks.append(chunk)
    return b"".join(chunks)


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            out.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            out.append(h)
    return out


def _pick_sheet(wb, sheet_sel: str | None):
    """Return (worksheet, sheet_name). sheet_sel: name or 1-based index string."""
    if sheet_sel is None:
        ws = wb.worksheets[0]
        return ws, ws.title
    if sheet_sel.isdigit():
        idx = int(sheet_sel)
        if not (1 <= idx <= len(wb.worksheets)):
            raise HTTPException(422, f"Sheet index {idx} out of range (1-{len(wb.worksheets)})")
        ws = wb.worksheets[idx - 1]
        return ws, ws.title
    if sheet_sel not in wb.sheetnames:
        raise HTTPException(422, f"Sheet '{sheet_sel}' not found. Available: {wb.sheetnames}")
    return wb[sheet_sel], sheet_sel


def _parse_xlsx(data: bytes, sheet_sel: str | None, header_row: int, max_rows: int):
    import openpyxl

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(422, f"Not a valid .xlsx file: {e}")
    ws, sheet_name = _pick_sheet(wb, sheet_sel)

    rows = ws.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows, None)
    header = next(rows, None)
    if header is None:
        raise HTTPException(422, "Sheet is empty")

    headers = _dedupe_headers([str(h).strip() if h is not None else f"col_{i + 1}"
                               for i, h in enumerate(header)])
    data = []
    for row in rows:
        if len(data) >= max_rows:
            break
        if all(c is None for c in row):
            continue
        data.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    wb.close()
    return sheet_name, headers, data


def _parse_csv(data: bytes, header_row: int, max_rows: int):
    text = data.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = iter(csv.reader(io.StringIO(text), dialect))
    for _ in range(header_row - 1):
        next(rows, None)
    header = next(rows, None)
    if header is None:
        raise HTTPException(422, "File is empty")

    headers = _dedupe_headers([h.strip() if h else f"col_{i + 1}" for i, h in enumerate(header)])
    data = []
    for row in rows:
        if len(data) >= max_rows:
            break
        if not any(c.strip() for c in row):
            continue
        data.append({headers[i]: (row[i].strip() if i < len(row) else "") for i in range(len(headers))})
    return "csv", headers, data


def _detect_format(url: str, content_type: str | None) -> str:
    path = urllib.parse.urlparse(url).path.lower()
    if path.endswith(".xlsx"):
        return "xlsx"
    if path.endswith(".xls"):
        return "xls"
    if path.endswith(".tsv"):
        return "tsv"
    if path.endswith(".csv"):
        return "csv"
    if content_type:
        if "spreadsheetml" in content_type or "excel" in content_type:
            return "xlsx"
        if "csv" in content_type or "text/plain" in content_type:
            return "csv"
    return "unknown"


def _sniff_format(data: bytes) -> str:
    """Detect format from magic bytes when URL/content-type give no hint."""
    if data[:4] == b"PK\x03\x04":
        # zip container: check for xlsx signature inside
        try:
            import zipfile
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                names = z.namelist()
                if any(n.startswith("xl/") for n in names) or "[Content_Types].xml" in names:
                    ct = z.read("[Content_Types].xml") if "[Content_Types].xml" in names else b""
                    if b"spreadsheetml" in ct or any(n.startswith("xl/") for n in names):
                        return "xlsx"
        except Exception:
            pass
        return "unknown"  # zip but not xlsx (docx/pptx/plain zip)
    if data[:4] == b"\xd0\xcf\x11\xe0":
        return "xls"  # OLE2 compound document
    # text-ish: try UTF-8 decode; if it looks like delimited text -> csv
    try:
        sample = data[:8192].decode("utf-8-sig")
    except UnicodeDecodeError:
        return "unknown"
    if "\x00" in sample:
        return "unknown"
    try:
        csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return "csv"
    except csv.Error:
        return "unknown"


# ---------------------------------------------------------------- endpoints

def _convert(url: str, sheet: str | None, header_row: int, max_rows: int) -> dict:
    url = _check_url(url)
    try:
        r = requests.head(url, timeout=10, allow_redirects=True)
        content_type = r.headers.get("content-type") if r.status_code < 400 else None
    except requests.RequestException:
        content_type = None
    data = _download(url)
    fmt = _detect_format(url, content_type)
    if fmt == "unknown":
        fmt = _sniff_format(data)
        if fmt == "unknown":
            raise HTTPException(422, "Unsupported format: expected .xlsx/.csv/.tsv URL")

    if fmt == "xlsx":
        sheet_name, headers, rows = _parse_xlsx(data, sheet, header_row, max_rows)
    elif fmt in ("csv", "tsv"):
        sheet_name, headers, rows = _parse_csv(data, header_row, max_rows)
    elif fmt == "xls":
        raise HTTPException(422, ".xls (legacy Excel) not supported yet — convert to .xlsx first")
    else:
        raise HTTPException(422, "Unsupported format: expected .xlsx/.csv/.tsv URL")

    return {
        "source": url,
        "format": fmt,
        "sheet": sheet_name,
        "row_count": len(rows),
        "column_count": len(headers),
        "headers": headers,
        "data": rows,
        "truncated": len(rows) >= max_rows,
    }


@router.post("/convert")
def convert(req: ConvertRequest) -> dict:
    if not req.resolved_url():
        raise HTTPException(422, "Missing 'url' (or 'datasource')")
    return _convert(req.resolved_url(), req.sheet, req.header_row, req.max_rows)


@router.get("/convert")
def convert_get(
    url: str | None = Query(None, description="URL of the Excel/CSV file"),
    datasource: str | None = Query(None, description="Alias for url (mustache template field name)"),
    sheet: str | None = Query(None, description="Sheet name or 1-based index"),
    header_row: int = Query(1, ge=1, description="1-based header row"),
    max_rows: int = Query(DEFAULT_MAX_ROWS, ge=1, le=100000),
) -> dict:
    target = url or datasource
    if not target:
        raise HTTPException(422, "Missing 'url' (or 'datasource')")
    return _convert(target, sheet, header_row, max_rows)
